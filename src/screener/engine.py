import os
import sqlite3
import pandas as pd
import numpy as np

class NiftyScreenerEngine:
    def __init__(self, db_path="database/nifty100.db"):
        self.db_path = db_path
        self.df = self._load_data_layer()
        # Compute dynamic scores immediately upon loading data layer for downstream screening
        self._calculate_composite_quality_scores()

    def _load_data_layer(self):
        if not os.path.exists(self.db_path):
            raise FileNotFoundError(f"Database not detected at: {self.db_path}")
        
        conn = sqlite3.connect(self.db_path)
        df = pd.read_sql("SELECT * FROM financial_ratios", conn)
        conn.close()
        
        df.columns = [col.lower().strip() for col in df.columns]
        
        if 'year' in df.columns:
            df['clean_year'] = df['year'].astype(str).str.extract(r'(\d{4})').astype(float)
        else:
            df['clean_year'] = 2024.0

        # Base Metric Alias Handlers
        alias_dict = {
            'pe': ['pe_ratio', 'p_e', 'pe', 'p/e'],
            'pb': ['pb_ratio', 'p_b', 'pb', 'p/b'],
            'de': ['debt_to_equity', 'd_e', 'de', 'd/e'],
            'roe': ['return_on_equity_pct', 'roe', 'return_on_equity'],
            'roce': ['return_on_capital_employed_pct', 'roce', 'return_on_capital_employed'],
            'npm': ['net_profit_margin_pct', 'npm', 'net_profit_margin'],
            'fcf': ['free_cash_flow_cr', 'fcf', 'free_cash_flow'],
            'rev_cagr': ['revenue_cagr_5yr', 'revenue_cagr', 'revenue_cagr_10yr'],
            'pat_cagr': ['pat_cagr_5yr', 'pat_cagr', 'pat_cagr_10yr'],
            'div_yield': ['dividend_yield_pct', 'dividend_yield', 'div_yield']
        }
        
        for target, aliases in alias_dict.items():
            if target not in df.columns:
                for alias in aliases:
                    if alias in df.columns:
                        df[target] = df[alias]
                        break
                if target not in df.columns:
                    df[target] = 0.0

        numeric_targets = ['roe', 'roce', 'npm', 'de', 'fcf', 'rev_cagr', 'pat_cagr', 'pe', 'pb', 'div_yield']
        for col in numeric_targets:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
            
        return df

    def _get_sector_column(self, df):
        possible_names = ['broad_sector', 'sector', 'industry', 'sector_name']
        for name in possible_names:
            if name in df.columns:
                return name
        return None

    def _winsorize_and_scale(self, series):
        """Day 17 Component: P10/P90 Winsorization and 0-100 scaling transformation"""
        if series.empty or series.nunique() <= 1:
            return pd.Series(50.0, index=series.index)
        
        p10 = np.percentile(series, 10)
        p90 = np.percentile(series, 90)
        
        # Cap extremes at 10th and 90th percentiles
        capped = np.clip(series, p10, p90)
        
        # Scale to 0-100
        min_v, max_v = capped.min(), capped.max()
        if max_v == min_v:
            return pd.Series(50.0, index=series.index)
            
        return ((capped - min_v) / (max_v - min_v)) * 100.0

    def _calculate_composite_quality_scores(self):
        """Day 17 Component: Computational Weighting Matrix Strategy"""
        df = self.df
        sector_col = self._get_sector_column(df)
        
        # 1. Setup proxy markers for calculation components missing in direct raw tables
        df['fcf_positive_flag'] = (df['fcf'] > 0).astype(float) * 100.0
        df['cfo_pat_ratio'] = df['fcf'] # standard data layer fallback proxy
        df['icr_score'] = 100.0 - np.clip(df['de'] * 20, 0, 100) # Proxy inverse logic for leverage stability

        # 2. Winsorize and Scale metrics relative to broad sector peers
        metrics_to_scale = ['roe', 'roce', 'npm', 'rev_cagr', 'pat_cagr', 'fcf_positive_flag', 'cfo_pat_ratio', 'de', 'icr_score']
        
        scaled_features = pd.DataFrame(index=df.index)
        
        if sector_col:
            for metric in metrics_to_scale:
                # Groupby sector relative processing
                if metric == 'de': # lower is better transformation
                    scaled_features[f's_{metric}'] = df.groupby(sector_col)[metric].transform(lambda x: 100.0 - self._winsorize_and_scale(x))
                else:
                    scaled_features[f's_{metric}'] = df.groupby(sector_col)[metric].transform(self._winsorize_and_scale)
        else:
            for metric in metrics_to_scale:
                if metric == 'de':
                    scaled_features[f's_{metric}'] = 100.0 - self._winsorize_and_scale(df[metric])
                else:
                    scaled_features[f's_{metric}'] = self._winsorize_and_scale(df[metric])

        # 3. Apply Explicit Structural Dimension Weights Mapping
        # 35% Profitability
        profitability_score = (scaled_features['s_roe'] * 0.15) + (scaled_features['s_roce'] * 0.10) + (scaled_features['s_npm'] * 0.10)
        # 30% Cash Quality
        cash_quality_score = (scaled_features['s_pat_cagr'] * 0.15) + (scaled_features['s_cfo_pat_ratio'] * 0.10) + (scaled_features['s_fcf_positive_flag'] * 0.05)
        # 20% Growth
        growth_score = (scaled_features['s_rev_cagr'] * 0.10) + (scaled_features['s_pat_cagr'] * 0.10)
        # 15% Leverage
        leverage_score = (scaled_features['s_de'] * 0.10) + (scaled_features['s_icr_score'] * 0.05)

        # Final Composite Aggregation
        df['composite_quality_score'] = (profitability_score + cash_quality_score + growth_score + leverage_score) / 0.70
        df['composite_quality_score'] = df['composite_quality_score'].clip(0.0, 100.0).round(2)

    def filter_by_custom_bounds(self, filter_params: dict):
        if self.df.empty:
            return pd.DataFrame()

        latest_year = self.df['clean_year'].max()
        df_latest = self.df[self.df['clean_year'] == latest_year].copy()
        sector_col = self._get_sector_column(df_latest)

        for metric, threshold in filter_params.items():
            key = metric
            if 'roe' in metric: key = 'roe'
            elif 'debt' in metric or 'de' in metric: key = 'de'
            
            if key in df_latest.columns:
                if key == 'de':
                    if sector_col:
                        is_financial = df_latest[sector_col].str.lower().str.contains('financial', na=False)
                        df_latest = df_latest[(df_latest[key] <= threshold) | is_financial]
                    else:
                        df_latest = df_latest[df_latest[key] <= threshold]
                else:
                    df_latest = df_latest[df_latest[key] >= threshold]

        # Always sort outputs by composite quality score descending for Day 17 output mandate
        if 'composite_quality_score' in df_latest.columns:
            df_latest = df_latest.sort_values(by='composite_quality_score', ascending=False)

        return df_latest

    def run_preset_screener_day16(self, preset_name: str):
        latest_year = self.df['clean_year'].max()
        df_latest = self.df[self.df['clean_year'] == latest_year].copy()
        
        sector_col = self._get_sector_column(df_latest)
        is_financial = df_latest[sector_col].str.lower().str.contains('financial', na=False) if sector_col else pd.Series(False, index=df_latest.index)

        if preset_name == "Quality Compounder":
            cond = (df_latest['roe'] > 15.0) & ((df_latest['de'] < 1.0) | is_financial) & (df_latest['fcf'] > 0) & (df_latest['rev_cagr'] > 10.0)
            res = df_latest[cond]
        elif preset_name == "Value Pick":
            cond = (df_latest['pe'] > 0) & (df_latest['pe'] < 30.0) & (df_latest['pb'] < 4.0)
            res = df_latest[cond]
        elif preset_name == "Growth Accelerator":
            cond = (df_latest['pat_cagr'] > 20.0) & (df_latest['rev_cagr'] > 15.0) & ((df_latest['de'] < 2.0) | is_financial)
            res = df_latest[cond]
        elif preset_name == "Dividend Champion":
            cond = (df_latest['div_yield'] > 1.5) & (df_latest['fcf'] > 0)
            res = df_latest[cond]
        elif preset_name == "Debt-Free Blue Chip":
            cond = (df_latest['de'] == 0) & (df_latest['roe'] > 12.0)
            res = df_latest[cond]
        elif preset_name == "Turnaround Watch":
            cond = (df_latest['rev_cagr'] > 10.0) & (df_latest['fcf'] > 0)
            res = df_latest[cond]
        else:
            return pd.DataFrame()

        if 'composite_quality_score' in res.columns:
            res = res.sort_values(by='composite_quality_score', ascending=False)
        return res

    def export_all_presets_to_excel(self, output_path="output/screener_output.xlsx"):
        """Day 17 Component: Generates multi-sheet Excel output with conditional color-coding"""
        from openpyxl.styles import PatternFill
        
        # Ensure 'output/' directory structure exists safely
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        presets = ["Quality Compounder", "Value Pick", "Growth Accelerator", "Dividend Champion", "Debt-Free Blue Chip", "Turnaround Watch"]
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for p_name in presets:
                df_res = self.run_preset_screener_day16(p_name)
                
                # Excel tab character limit safety rule (Max 31)
                sheet_label = p_name[:31]
                df_res.to_excel(writer, sheet_name=sheet_label, index=False)
                
                worksheet = writer.sheets[sheet_label]
                
                # Custom Soft Palette Fills
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Map headers dynamically to target correct column locations
                col_headers = [cell.value for cell in worksheet[1]]
                roe_idx = col_headers.index('roe') + 1 if 'roe' in col_headers else None
                de_idx = col_headers.index('de') + 1 if 'de' in col_headers else None
                
                # Inject styling row-by-row matching conditional presets
                for row in range(2, worksheet.max_row + 1):
                    # 1. ROE formatting check
                    if roe_idx:
                        cell = worksheet.cell(row=row, column=roe_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.fill = green_fill if cell.value >= 15.0 else red_fill
                    
                    # 2. D/E formatting check
                    if de_idx:
                        cell = worksheet.cell(row=row, column=de_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.fill = green_fill if cell.value < 1.0 else red_fill

        print(f"📊 Excel Generation: Successfully saved output sheets to '{output_path}' with color codes applied!")

if __name__ == "__main__":
    engine = NiftyScreenerEngine()
    
    print("\n" + "="*50)
    print(" 🚀 EXECUTION VERIFICATION: DAY 17 COMPOSITE SCORING ENGINE ")
    print("="*50)
    
    # Check random samples for scores
    latest_year = engine.df['clean_year'].max()
    sample_view = engine.df[engine.df['clean_year'] == latest_year][['company_id', 'roe', 'de', 'composite_quality_score']].head(5)
    print("📋 Sample Evaluated Entities with Winsorized Quality Scores:")
    print(sample_view.to_string(index=False))

    # Run the dynamic Excel exporter with color formatting
    engine.export_all_presets_to_excel()