import os
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class PeerComparisonReporter:
    def __init__(self, db_path="database/nifty100.db", output_path="output/peer_comparison.xlsx"):
        self.db_path = db_path
        self.output_path = output_path
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self.conn = sqlite3.connect(self.db_path)
        
        # Color palettes optimized for high corporate readability
        self.fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green >= 75
        self.fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow 25-75
        self.fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")    # Soft red <= 25
        self.fill_gold = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")   # Gold/Amber Benchmark
        self.fill_median = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Group Median Base

    def load_data(self):
        df_ratios = pd.read_sql("SELECT * FROM financial_ratios", self.conn)
        df_ratios.columns = [col.lower().strip() for col in df_ratios.columns]
        
        df_ranks = pd.read_sql("SELECT * FROM peer_percentiles", self.conn)
        
        try:
            df_meta = pd.read_sql("SELECT company_id, company_name FROM companies", self.conn)
            df_meta.columns = [col.lower().strip() for col in df_meta.columns]
        except:
            df_meta = pd.DataFrame({'company_id': df_ratios['company_id'].unique()})
            df_meta['company_name'] = df_meta['company_id'] + " Limited"
            
        return df_ratios, df_ranks, df_meta

    def build_excel_report(self):
        df_ratios, df_ranks, df_meta = self.load_data()
        
        if df_ranks.empty:
            print("❌ Failure context: Base ranks empty.")
            return

        peer_groups = df_ranks['peer_group_name'].unique()
        wb = Workbook()
        wb.remove(wb.active)
        
        exclude_cols = {'company_id', 'company_name', 'year', 'clean_year', 'sector', 'industry', 'broad_sector'}
        metric_cols = [col for col in df_ratios.columns if col not in exclude_cols]

        for group in peer_groups:
            sheet_title = str(group)[:31].replace(":", "").replace("/", "")
            ws = wb.create_sheet(title=sheet_title)
            
            group_companies = df_ranks[df_ranks['peer_group_name'] == group]['company_id'].unique()
            if len(group_companies) == 0:
                continue
                
            df_group_ratios = df_ratios[df_ratios['company_id'].isin(group_companies)].copy()
            df_group_ratios = pd.merge(df_group_ratios, df_meta, on='company_id', how='left')
            
            df_group_ranks = df_ranks[df_ranks['peer_group_name'] == group]
            pivot_ranks = df_group_ranks.pivot_table(index='company_id', columns='metric', values='percentile_rank', aggfunc='last').reset_index()

            headers = ['Company ID', 'Company Name']
            actual_metrics = [m for m in metric_cols if m in df_group_ratios.columns]
            
            # Ensure 20+ columns are dynamically handled
            for col in actual_metrics:
                headers.append(f"{col.upper()}")
                headers.append(f"{col.upper()} (% Rank)")
                
            ws.append(headers)
            
            row_idx = 2
            for _, row in df_group_ratios.iterrows():
                comp_id = row['company_id']
                comp_name = row.get('company_name', f"{comp_id} Corp")
                
                row_data = [comp_id, comp_name]
                comp_rank_row = pivot_ranks[pivot_ranks['company_id'] == comp_id]
                
                for col in actual_metrics:
                    val = pd.to_numeric(row[col], errors='coerce')
                    val = float(val) if not np.isnan(val) else 0.0
                    
                    rank_val = 50.0
                    if not comp_rank_row.empty and col in comp_rank_row.columns:
                        rank_val = float(comp_rank_row.iloc[0][col])
                        
                    row_data.append(val)
                    row_data.append(rank_val)
                    
                ws.append(row_data)
                
                # Rule: First element row tagged as explicit benchmark row context
                is_benchmark = (row_idx == 2)
                
                for c_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=c_idx)
                    
                    if is_benchmark:
                        cell.fill = self.fill_gold  # Benchmark row gets strict style priority
                    elif c_idx > 2 and c_idx % 2 == 0:
                        # Color coding applies to non-benchmark ranks safely
                        cell_val = cell.value
                        if cell_val is not None:
                            if cell_val >= 75.0:
                                cell.fill = self.fill_green
                            elif cell_val <= 25.0:
                                cell.fill = self.fill_red
                            else:
                                cell.fill = self.fill_yellow
                row_idx += 1
                
            # Bottom Summary Row Matrix Calculation
            median_row_data = ['GROUP MEDIAN', '-']
            for col in actual_metrics:
                val_median = df_group_ratios[col].median()
                val_median = float(val_median) if not np.isnan(val_median) else 0.0
                
                if col in pivot_ranks.columns:
                    rank_median = pivot_ranks[col].median()
                else:
                    rank_median = 50.0
                    
                median_row_data.append(val_median)
                median_row_data.append(rank_median)
                
            ws.append(median_row_data)
            
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c_idx)
                cell.fill = self.fill_median
                cell.font = Font(bold=True)
                
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

        wb.save(self.output_path)
        print(f"✅ Excel Verification Layer: Successfully exported to '{self.output_path}'")

    def close(self):
        self.conn.close()

if __name__ == "__main__":
    reporter = PeerComparisonReporter()
    reporter.build_excel_report()
    reporter.close()